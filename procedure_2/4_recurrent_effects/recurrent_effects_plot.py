"""
Plot the 'Evoked response' stem figure from recurrent_effects_plot.xlsx (sheet 'Fig 2d', A1:K2).

Usage:
    python plot_evoked_response.py [path/to/recurrent_effects_plot.xlsx]

Requires: openpyxl, matplotlib
"""
import sys
import openpyxl
import matplotlib.pyplot as plt

# ---- Load data (row 2 = normalized firing values, A2:K2) --------------------
xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "recurrent_effects_plot.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb["Fig 2d"]
labels = [c.value for c in ws[1][1:11]]   # B1:K1  -> Pulse 1, Pulse 2_1, ...
vals   = [c.value for c in ws[2][1:11]]   # B2:K2  -> firing amplitudes

# ---- X positions ------------------------------------------------------------
# The xlsx stores only amplitudes, not timestamps. The 10 values are grouped
# into 4 pulse trains (1, 2, 3, 4 pulses). Trains are centered at ~0/2.3/4.4/6.5 s
# with 0.2 s spacing within a train, to match the reference figure layout.
xpos = [0,                       # Pulse 1
        2.2, 2.4,                # Pulse 2
        4.2, 4.4, 4.6,           # Pulse 3
        6.2, 6.4, 6.6, 6.8]      # Pulse 4

# ---- Plot -------------------------------------------------------------------
plt.rcParams.update({"font.size": 15, "font.family": "DejaVu Sans"})
fig, ax = plt.subplots(figsize=(5.2, 2.6))

green = "#2ca02c"
markers, stems, base = ax.stem(xpos, vals, linefmt="-", markerfmt="s", basefmt=" ")
plt.setp(stems, color=green, linewidth=1.3)
plt.setp(markers, color=green, markersize=6, markeredgecolor=green)

ax.set_xlabel("Time (s)")
ax.set_ylabel("Normalized firing")
ax.set_xlim(-0.7, 7.3)
ax.set_ylim(0, 0.45)
ax.set_xticks([0, 2, 4, 6])
ax.set_yticks([0, 0.2, 0.4])
ax.text(0.04, 0.93, "Evoked response", transform=ax.transAxes,
        va="top", ha="left", fontsize=15)
ax.spines[["top", "right"]].set_visible(False)
ax.tick_params(direction="out")

plt.tight_layout()
plt.savefig("evoked_response_fig.png", dpi=200, bbox_inches="tight")
plt.show()
